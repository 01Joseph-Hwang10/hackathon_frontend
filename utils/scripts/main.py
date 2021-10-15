from typing import List
from constants import ROOT
import os
from openpyxl import load_workbook
from openpyxl.cell import Cell
import re
import json

data_root = os.path.join(ROOT, 'data')
source = os.path.join(data_root, 'question-weights.backup.ts')
target = os.path.join(data_root, 'question-weights.ts')
weights = os.path.join(data_root, 'question_weight_edited.xlsx')

with open(source, 'r') as rf:
    regex = re.compile('export default (.*)', re.DOTALL)
    raw = regex.search(rf.read()).group(1)
    data = json.loads(raw)

wb = load_workbook(weights)

numbers = [2, 1, 3, 8, 7, 4, 6, 9, 5]
assert(len(numbers) == 9)
number_idx = [ i - 1 for i in numbers]

for i in range(len(number_idx)):
    choices = []
    cols_iterator = wb[wb.sheetnames[i]].columns
    cols: List[List[Cell]] = []
    titles: List[Cell] = list(next(cols_iterator))
    titles.pop(0)
    while True:
        try:
            col: List[Cell] = list(next(cols_iterator))
            col.pop(0)
            cols.append(col)
        except:
            break
    try:
        assert(len(cols) == len(data[number_idx[i]]['choices']))
    except:
        print("#%d : %s" % (i + 1, data[number_idx[i]]['title']))
        break
    for j in range(len(cols)):
        choice = {
            "title": data[number_idx[i]]['choices'][j]["title"]
        }
        vars = []
        for k in range(len(titles)):
            var = {
                "title": titles[k].value,
                "tagTitle": [],
                "weight": str(cols[j][k].value)
            }
            vars.append(var)
        choice["vars"] = vars
        choices.append(choice)
    data[number_idx[i]]["choices"] = choices

with open(target, 'w') as wf:
    wf.write('/* eslint-disable */\n')
    wf.write('export default ')
    wf.write(json.dumps(data, indent=2))

print('Editing successful!')
